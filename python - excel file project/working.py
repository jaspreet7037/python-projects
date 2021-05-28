from openpyxl import workbook , load_workbook # i have imported workbook and load_workbook from openpyxl module
from openpyxl.utils import get_column_letter

wb = load_workbook('C:\\Users\\JASPREET SINGH\\OneDrive\\Desktop\\python projects\\python - excel file project\\grades.xlsx')
ws = wb.active
"""ws['A2'].value = 'jaspreet'
ws['A3'].value = 'aakarsh'
ws['A4'].value = 'tanmay'

ws.unmerge_cells("A1:D1")"""

for row in range(1,11):
    for col in range(1,6):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)

wb.save('C:\\Users\\JASPREET SINGH\\OneDrive\\Desktop\\python projects\\python - excel file project\\grades.xlsx')

#ws.append("")