from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 

data = {
        "Aakrosh Tiwari": {
                          "Maths":95,
                          "Physics":92,
                          "Chemistry":91,
                          "English":90,
                          "Hindi":89,
                          "SUPW":100
        },
        "Aatika faisal": {
                         "Maths":95,
                         "Physics":90,
                         "chemistry":91,
                         "English":89,
                         "Hindi":80,
                         "SUPW":100
        },
        "Abhay Mishra": {
                        "Maths":95,
                        "Physics":90,
                        "Chemistry":89,
                        "English":85,
                        "Computer":94,
                        "SUPW":100
        },
        "Abhinav kumar": {
                         "Maths":95,
                         "Physics":93,
                         "Chemistry":91,
                         "English":90,
                         "Hindi":80,
                         "SUPW":100
        },
        "Aditya Pandey": {
                         "Maths":95,
                         "Physics":92,
                         "Chemistry":91,
                         "English":87,
                         "Hindi":80,
                         "SUPW":100
        },
        "Agamya Awasthi": {
                          "Biology":95,
                          "physics":92,
                          "Chemistry":96,
                          "English":85,
                          "Hindi":86,
                          "SUPW":100
        },
        "Amolika Soni": {
                        "Maths":95,
                        "Physics":87,
                        "Chemistry":89,
                        "Hindi":78,
                        "SUPW":100
        },
        "Ananaya Singh": {
                         "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Anubhav Singh": { 
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Anuj Chauhan": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Anuraj Singh Yadav": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Arpita Singh": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Aryan Verma": {
                       "Maths":95,
                       "physics":92,
                       "Chemistry":96,
                       "English":85,
                       "Computer":97,
                       "SUPW":100
        },
        "Ashish Kumar": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Atharva Trivedi": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Aviral Vishwakarma": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Ayush Dubey": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Ayush Singh": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Divyanjali": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Computer":97,
                         "SUPW":100
             
        },
        "Divyanshi Bhadauria": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Computer":90,
                         "SUPW":100
        },
        "Diya Shukla": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Eram Saher Fatima": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Hardik Srivastava": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Harsh Bhardwaj": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100
        },
        "Harshita Singh": {
                "Maths":95,
                         "Physics":90,
                         "Chemistry":92,
                         "Hindi":80,
                         "SUPW":100       
        },
        "Ishita Pandey": {
                         "Biology":95,
                          "physics":92,
                          "Chemistry":96,
                          "English":85,
                          "Hindi":86,
                          "SUPW":100
        },
        "Jaspreet Singh Virk": {
                               "Maths":95,
                               "physics":92,
                               "Chemistry":96,
                               "English":90,
                               "Computer":98,
                               "SUPW":100
        },
        "Jiya Singh": {
                      "Maths":95,
                      "physics":92,
                      "Chemistry":96,
                      "English":85,
                      "Hindi":86,
                      "SUPW":100
        },
        "Kalpana Pal": {
                       "Maths":95,
                       "physics":92,
                       "Chemistry":96,
                       "English":85,
                       "Hindi":86,
                       "SUPW":100
        },
        "Kartikey Chandra": {
                            "Maths":95,
                            "physics":92,
                            "Chemistry":96,
                            "English":85,
                            "Hindi":86,
                            "SUPW":100
        },
        "Khushi Katiyar": {
                          "Maths":95,
                          "physics":92,
                          "Chemistry":96,
                          "English":85,
                          "Hindi":86,
                          "SUPW":100
        },
        "Mahek Shukla": {
                        "Maths":95,
                        "physics":92,
                        "Chemistry":96,
                        "English":85,
                        "Hindi":86,
                        "SUPW":100
        },
        "Mayank Vishwakarma": {
                              "Maths":95,
                              "physics":92,
                              "Chemistry":96,
                              "English":85,
                              "Computer":86,
                              "SUPW":100
        },
        "Moubani Das": {
                       "Biology":95,
                       "physics":92,
                       "Chemistry":96,
                       "English":85,
                       "Hindi":86,
                       "SUPW":100
        },
        "Pawani Shukla": {
                         "Maths":95,
                         "physics":92,
                         "Chemistry":96,
                         "English":85,
                         "Hindi":86,
                         "SUPW":100
        },
        "Prakhar Tiwari": {
                          "Maths":95,
                          "physics":92,
                          "Chemistry":96,
                          "English":85,
                          "Hindi":86,
                          "SUPW":100
        },
        "Prashant Kumar Pandey": {
                                 "Maths":95,
                                 "physics":92,
                                 "Chemistry":96,
                                 "English":83,
                                 "Hindi":86,
                                 "SUPW":100
        },
        "Pratham Khanna": {
                          "Maths":95,
                          "physics":92,
                          "Chemistry":96,
                          "English":85,
                          "Computer":100,
                          "SUPW":100
        },
        "Priyani Garg": {
                        "Maths":95,
                        "physics":92,
                        "Chemistry":96,
                        "English":85,
                        "Computer":95,
                        "SUPW":100
        },
        "Rashmi": {
                  "Maths":95,
                  "physics":92,
                  "Chemistry":96,
                  "English":85,
                  "Hindi":86,
                  "SUPW":100
        },
        "Riya Dey": {
                    "Maths":95,
                    "physics":92,
                    "Chemistry":96,
                    "English":85,
                    "Hindi":86,
                    "SUPW":100
        },
        "Sachin Singh Rana": {
                             "Maths":95,
                             "physics":92,
                             "Chemistry":96,
                             "English":85,
                             "Hindi":86,
                             "SUPW":100
        },
        "Sanchit Singh": {
                         "Maths":95,
                         "physics":92,
                         "Chemistry":96,
                         "English":85,
                         "Hindi":86,
                         "SUPW":100
        },
        "Sanjana Pal": {
                       "Maths":95,
                       "physics":92,
                       "Chemistry":96,
                       "English":85,
                       "Hindi":86,
                       "SUPW":100
        },
        "Sarthak Omar": {
                        "Maths":95,
                        "physics":92,
                        "Chemistry":96,
                        "English":85,
                        "Hindi":86,
                        "SUPW":100
        },
        "Shantanu Agnihotri": {
                              "Maths":95,
                              "physics":92,
                              "Chemistry":96,
                              "English":85,
                              "Hindi":86,
                              "SUPW":100
        },
        "Shek Mohammad Ajmal": {
                               "Maths":95,
                               "physics":92,
                               "Chemistry":96,
                               "English":85,
                               "Hindi":86,
                               "SUPW":100
        },
        "Shristi Goswami": {
                           "Maths":95,
                           "physics":92,
                           "Chemistry":96,
                           "English":85,
                           "Hindi":86,
                           "SUPW":100
        },
        "Shristi Singh": {
                        "Maths":95,
                        "physics":92,
                        "Chemistry":96,
                        "English":85,
                        "Hindi":86,
                        "SUPW":100
        },
        "Shubham Kumar Verma": {
                               "Maths":95,
                               "physics":92,
                               "Chemistry":96,
                               "English":85,
                               "Hindi":86,
                               "SUPW":100
        },
        "Shubhanshi": {
                      "Maths":95,
                      "physics":92,
                      "Chemistry":96,
                      "English":85,
                      "Hindi":86,
                      "SUPW":100
        },
        "Suryanshi Singh": {
                           "Maths":95,
                           "physics":92,
                           "Chemistry":96,
                           "English":85,
                           "Hindi":86,
                           "SUPW":100
        },
        "Tanishq Gautam": {
                          "Maths":95,
                          "physics":92,
                          "Chemistry":96,
                          "English":85,
                          "Hindi":86,
                          "SUPW":100
        },
        "Tanishq Mishra": {
                          "Maths":95,
                          "physics":92,
                          "Chemistry":96,
                          "English":85,
                          "Hindi":88,
                          "SUPW":100
        },
        "Tarnija Chaudary": {
                            "Biology":95,
                            "physics":92,
                            "Chemistry":96,
                            "English":85,
                            "Hindi":90,
                            "SUPW":100                      
        },
        "Tarnika Chaudary" : {
                             "Biology":95,
                             "physics":92,
                             "Chemistry":96,
                             "English":85,
                             "Hindi":89,
                             "SUPW":100
        },
        "Tejas Khushwaha": {
                "Maths":95,
                        "physics":92,
                        "Chemistry":96,
                        "English":85,
                        "Hindi":81,
                        "SUPW":100
        },
        "Ujjwal Dixit": {
                        "Maths":95,
                        "physics":92,
                        "Chemistry":96,
                        "English":85,
                        "Hindi":83,
                        "SUPW":100
        },
        "Yuvraj Singh": {
                        "Maths":95,
                        "physics":92,
                        "Chemistry":96,
                        "English":85,
                        "Computer":84,
                        "SUPW":100
        }
        
}

wb = Workbook("C:\\Users\\JASPREET SINGH\\OneDrive\\Desktop\\python projects\\python - excel file project\\class progress.xlsx")
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Aakrosh tiwari'].keys())
ws.append(headings)

for person in data:
        grades = list(data[person].values())
        ws.append([person] + grades)

for col in range(2,len(data['Aakrosh tiwari']) + 2):
        char = get_column_letter(col)
        ws[char + "7"] = f"=sum({char +'2'}:{char + '6'})/{len(data)}"

wb.save("C:\\Users\\JASPREET SINGH\\OneDrive\\Desktop\\python projects\\python - excel file project\\class progress.xlsx")